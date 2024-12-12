import advertools as adv
import pandas as pd
import streamlit as st
import time
from openpyxl import load_workbook
from polyfuzz import PolyFuzz
from polyfuzz.models import RapidFuzz
import io

# Inizializzazione della sessione Streamlit per cache
if 'matched_results' not in st.session_state:
    st.session_state.matched_results = None
    st.session_state.legacy_crawl = None
    st.session_state.new_crawl = None
    st.session_state.legacy_url_parse = None
    st.session_state.new_url_parse = None

# Imposta il matcher e il modello
matcher = RapidFuzz(n_jobs=1)
model = PolyFuzz(matcher)

# Configurazione della pagina
st.set_page_config(page_title="Redirect URL Mapper • NUR® Digital Marketing", page_icon="./Nur-simbolo-1080x1080.png")
st.image("./logo_nur_vettoriale.svg", width=100)
st.markdown("# Redirect URL Mapper")
st.markdown("Strumento per la mappatura automatica dei reindirizzamenti, che confronta gli URL precedenti e quelli nuovi in base a vari elementi della pagina (percorsi, slug, titoli, H1 e H2).")

# Sezione per istruzioni e requisiti in blocchi espandibili
with st.expander("Istruzioni"):
    st.markdown("""
    ### Come utilizzare lo strumento
    1. **Preparazione dei dati**:
       - **Per file Excel**:
         - Esegui una scansione con Screaming Frog sia del sito attualmente live che del sito staging/nuovo
         - Per ottenere risultati migliori, utilizza le sitemap XML come fonte di crawling
         - Puoi utilizzare l'URL List Mode per crawlare solo specifiche sezioni del sito
       
       - **Per lista URL**:
         - Prepara due liste di URL (vecchi e nuovi)
         - Assicurati che gli URL siano completi e uno per riga
         - Puoi copiare e incollare direttamente da Excel o altri strumenti

    2. **Utilizzo dello strumento**:
       - Scegli il metodo di input preferito (File Excel o Lista URL)
       - Carica o incolla i tuoi URL
       - Utilizza i cursori nella sidebar per regolare le soglie di similarità
    """)

with st.expander("Requisiti e formato dei file"):
    st.markdown("""
    ### Requisiti dei dati
    1. **Per file Excel**:
       - È richiesta la colonna "Address" con gli URL completi
       - Le colonne "Title 1", "H1-1", "H2-1" sono opzionali ma consigliate per match più accurati
       
    2. **Per lista URL**:
       - URLs completi con http(s)://
       - Un URL per riga
       - Le intestazioni delle colonne verranno ignorate automaticamente

    3. **Note importanti**:
       - Assicurati che gli URL non contengano spazi o caratteri speciali
       - Per file Excel, verifica che non ci siano celle con formattazione speciale
    """)

# Configurazione soglie nella sidebar
st.sidebar.header("Configurazione Soglie di Similarità")
threshold_url = st.sidebar.slider("Soglia similarità URL", 0.0, 1.0, 0.65, 0.01)
threshold_slug = st.sidebar.slider("Soglia similarità Slug", 0.0, 1.0, 0.65, 0.01)
threshold_title = st.sidebar.slider("Soglia similarità Titoli", 0.0, 1.0, 0.70, 0.01)
threshold_h1 = st.sidebar.slider("Soglia similarità H1", 0.0, 1.0, 0.90, 0.01)
threshold_h2 = st.sidebar.slider("Soglia similarità H2", 0.0, 1.0, 0.90, 0.01)

# Funzione per convertire gli URL incollati in DataFrame
def urls_to_dataframe(urls_text):
    # Dividi il testo in righe e rimuovi righe vuote
    urls = [url.strip() for url in urls_text.split('\n') if url.strip()]
    
    # Rimuovi eventuali intestazioni
    if urls and any(url.lower().startswith(('url', 'http', 'www')) for url in urls):
        urls = [url for url in urls if not url.lower() == 'url']
    
    # Crea DataFrame
    df = pd.DataFrame({
        'Address': urls,
        'Title 1': [''] * len(urls),
        'H1-1': [''] * len(urls),
        'H2-1': [''] * len(urls)
    })
    return df

# Funzione per eseguire il match iniziale
def perform_initial_match(data_type, from_data, to_data):
    model.match(from_data, to_data)
    matches = model.get_matches()
    matches["Similarity"] = matches["Similarity"].round(3)
    matches = matches.sort_values('Similarity', ascending=False)
    return matches

# Funzione per filtrare i risultati
def filter_and_join_results(matches_df, threshold, legacy_data, new_data, match_type='url'):
    filtered_df = matches_df[matches_df['Similarity'] >= threshold].copy()
    
    if match_type in ['url', 'slug']:
        join_df = pd.merge(filtered_df, legacy_data, left_on='From', right_on='path')
        join_df_2 = pd.merge(join_df, new_data, left_on='To', right_on='path')
        join_df_2.rename(
            columns={'url_x': 'URL Legacy', 'url_y': 'URL Nuovo', 
                    'path_x': 'Percorso Legacy', 'path_y': 'Percorso Nuovo'},
            inplace=True)
        result_df = join_df_2[['From', 'To', 'Similarity', 
                              'Percorso Legacy', 'Percorso Nuovo', 
                              'URL Legacy', 'URL Nuovo']]
    else:  # title, h1, h2
        join_df = pd.merge(filtered_df, legacy_data, 
                          left_on='From', 
                          right_on=f'{match_type.capitalize()}-1' if match_type != 'title' else 'Title 1')
        join_df_2 = pd.merge(join_df, new_data, 
                            left_on='To', 
                            right_on=f'{match_type.capitalize()}-1' if match_type != 'title' else 'Title 1')
        join_df_2.rename(columns={'Address_x': 'URL Legacy', 'Address_y': 'URL Nuovo'}, inplace=True)
        result_df = join_df_2[['From', 'To', 'Similarity', 'URL Legacy', 'URL Nuovo']]
    
    return result_df.drop_duplicates()

# Funzione per l'analisi dei crawl da file Excel
def analyze_crawls(crawls):
    with st.spinner('Elaborazione dei crawl del sito in corso...'):
        progress_bar = st.progress(0)
        try:
            input_files = []
            for crawl_index, crawl in enumerate(crawls):
                wb = load_workbook(filename=crawl)
                sheet_name = wb.sheetnames
                input_files.append([crawl, sheet_name])
                
                # Leggi il file Excel e verifica la colonna Address
                df_test = pd.read_excel(crawl, sheet_name=sheet_name[0])
                if 'Address' not in df_test.columns:
                    st.error(f"Colonna 'Address' mancante nel file {crawl.name}")
                    return
                
                progress_bar.progress((crawl_index + 1) / len(crawls))
                time.sleep(0.01)

            # Carica i crawl
            required_columns = ['Address']
            optional_columns = ['Title 1', 'H1-1', 'H2-1']
            
            if st.session_state.legacy_crawl is None:
                legacy_df = pd.read_excel(input_files[0][0], sheet_name=input_files[0][1][0])
                new_df = pd.read_excel(input_files[1][0], sheet_name=input_files[1][1][0])
                
                # Aggiungi colonne opzionali mancanti con valori vuoti
                for col in optional_columns:
                    if col not in legacy_df.columns:
                        legacy_df[col] = ''
                    if col not in new_df.columns:
                        new_df[col] = ''
                
                st.session_state.legacy_crawl = legacy_df
                st.session_state.new_crawl = new_df
                
                process_urls()

        except Exception as e:
            st.error(f"Si è verificato un errore durante l'elaborazione: {str(e)}")
            reset_session_state()

# Funzione per processare gli URL
def process_urls():
    try:
        # Parse URL
        legacy_urls = st.session_state.legacy_crawl['Address'].tolist()
        new_urls = st.session_state.new_crawl['Address'].tolist()
        
        st.session_state.legacy_url_parse = adv.url_to_df(legacy_urls)[['url', 'path', 'last_dir']]
        st.session_state.new_url_parse = adv.url_to_df(new_urls)[['url', 'path', 'last_dir']]
        
        # Esegui i match iniziali
        st.session_state.matched_results = {
            'url': perform_initial_match(
                'url',
                st.session_state.legacy_url_parse['path'],
                st.session_state.new_url_parse['path']
            ),
            'slug': perform_initial_match(
                'slug',
                st.session_state.legacy_url_parse['last_dir'],
                st.session_state.new_url_parse['last_dir']
            ),
            'title': perform_initial_match(
                'title',
                st.session_state.legacy_crawl['Title 1'],
                st.session_state.new_crawl['Title 1']
            ),
            'h1': perform_initial_match(
                'h1',
                st.session_state.legacy_crawl['H1-1'],
                st.session_state.new_crawl['H1-1']
            ),
            'h2': perform_initial_match(
                'h2',
                st.session_state.legacy_crawl['H2-1'],
                st.session_state.new_crawl['H2-1']
            )
        }
        
        # Filtra e prepara i risultati
        filtered_results = [
            filter_and_join_results(
                st.session_state.matched_results['url'],
                threshold_url,
                st.session_state.legacy_url_parse,
                st.session_state.new_url_parse,
                'url'
            ),
            filter_and_join_results(
                st.session_state.matched_results['slug'],
                threshold_slug,
                st.session_state.legacy_url_parse,
                st.session_state.new_url_parse,
                'slug'
            ),
            filter_and_join_results(
                st.session_state.matched_results['title'],
                threshold_title,
                st.session_state.legacy_crawl,
                st.session_state.new_crawl,
                'title'
            ),
            filter_and_join_results(
                st.session_state.matched_results['h1'],
                threshold_h1,
                st.session_state.legacy_crawl,
                st.session_state.new_crawl,
                'h1'
            ),
            filter_and_join_results(
                st.session_state.matched_results['h2'],
                threshold_h2,
                st.session_state.legacy_crawl,
                st.session_state.new_crawl,
                'h2'
            )
        ]
        
        display_results(filtered_results)
        
    except Exception as e:
        st.error(f"Si è verificato un errore durante l'elaborazione: {str(e)}")
        reset_session_state()

# Funzione per visualizzare i risultati
def display_results(match_dfs):
    sheet_names = ['URL Match', 'Slug Match', 'Title Match', 'H1 Match', 'H2 Match']
    
    # Stile CSS per le metriche
    st.markdown("""
        <style>
        .stats-card {
            background-color: #1E1E1E;
            border-radius: 10px;
            padding: 20px;
            margin: 10px 0;
        }
        .metric-title {
            color: #9E9E9E;
            font-size: 0.9em;
            margin-bottom: 5px;
        }
        .metric-value {
            color: #FFFFFF;
            font-size: 1.8em;
            font-weight: bold;
            margin: 0;
        }
        </style>
    """, unsafe_allow_html=True)

    # Per ogni tipo di match
    for i, df in enumerate(match_dfs):
        st.markdown(f"### {sheet_names[i]}")
        
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"""
                <div class="stats-card">
                    <div class="metric-title">Match trovati</div>
                    <div class="metric-value">{len(df):,}</div>
                </div>
            """, unsafe_allow_html=True)
        
        with col2:
            avg_similarity = df['Similarity'].mean() if len(df) > 0 else 0
            st.markdown(f"""
                <div class="stats-card">
                    <div class="metric-title">Similarità media</div>
                    <div class="metric-value">{avg_similarity:.1%}</div>
                </div>
            """, unsafe_allow_html=True)

    # Visualizzazione tabella e export
    selected_sheet = st.selectbox("Seleziona il match da visualizzare", sheet_names)
    sheet_index = sheet_names.index(selected_sheet)
    st.dataframe(match_dfs[sheet_index])

    # Preparazione file Excel per il download
    with pd.ExcelWriter('mappatura_url.xlsx') as writer:
        for df, sheet_name in zip(match_dfs, sheet_names):
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    with open("mappatura_url.xlsx", "rb") as file:
        st.download_button(
            label='Scarica l\'analisi del match',
            data=file,
            file_name='mappatura_url.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

# Funzione per resettare lo stato della sessione
def reset_session_state():
    st.session_state.matched_results = None
    st.session_state.legacy_crawl = None
    st.session_state.new_crawl = None
    st.session_state.legacy_url_parse = None
    st.session_state.new_url_parse = None

# Interfaccia principale con tabs
tab1, tab2 = st.tabs(["📁 Carica file Excel", "📝 Incolla URLs"])

with tab1:
    st.markdown("### Carica i file Excel")
    st.markdown("I file Excel devono contenere almeno la colonna 'Address' con gli URL completi. Le colonne 'Title 1', 'H1-1', 'H2-1' sono opzionali ma consigliate per match più accurati.")
    
    legacy_file = st.file_uploader('File degli URLs attualmente live', type='xlsx', key='legacy')
    if legacy_file is not None:
        new_file = st.file_uploader('File degli URLs staging', type='xlsx', key='new')
        if new_file is not None:
            if st.button("Analizza files"):
                analyze_crawls([legacy_file, new_file])

with tab2:
    st.markdown("### Incolla gli URLs")
    st.markdown("Inserisci gli URL completi, uno per riga.")
    
    legacy_urls = st.text_area(
        "URLs attualmente live",
        height=200,
        help="Incolla qui gli URL vecchi, uno per riga"
    )
    new_urls = st.text_area(
        "URLs staging",
        height=200,
        help="Incolla qui gli URL nuovi, uno per riga"
    )
    
    if st.button("Analizza URLs") and legacy_urls and new_urls:
        # Converti gli URL in DataFrame
        st.session_state.legacy_crawl = urls_to_dataframe(legacy_urls)
        st.session_state.new_crawl = urls_to_dataframe(new_urls)
        process_urls()

# Pulsante per resettare la cache
if st.button('Resetta cache e ricarica'):
    reset_session_state()
    st.rerun()

# Footer
st.markdown("---")
st.markdown("© 2024 [NUR® Digital Marketing](https://www.nur.it/). Tutti i diritti riservati.")
